import openpyxl
import requests #for downloading images
from bs4 import BeautifulSoup as BS

loc = ("excel_datasheet.xlsx")

wb = openpyxl.load_workbook(loc, read_only=True)
sheet = wb.active

def getdata(url):
    r = requests.get(url)
    return r.text

address = "https://www.food.com/recipe/"
data = []
valid = 0
stop = False
for i in range(2, 5000): #iterating through each row

    try:

        #try to get the main image using the image ID in column 1
        htmldata = getdata(address+str(sheet.cell(row=i, column=1).value))
        soup = BS(htmldata, 'html.parser')
        # for item in soup.find_all('img'):
        #     print(item['src'])
        s = soup.find('div', class_='recipe-default-image recipe-hero__item theme-gradient svelte-jlgald')
        # print(s)
        # break
        if(s == None):
            continue

        tmp_soup = s.find('img')

        # print(tmp_soup)
        # break
        img_url = tmp_soup['src']

        #request for image using url
        with open(str(sheet.cell(row=i, column=1).value)+'.jpg', 'wb') as handle:
            response = requests.get(img_url, stream=True)
            for block in response.iter_content(1024):
                if not block:
                    break

                handle.write(block)

        # #add data to main test data list
        # tmp = []
        # tmp1 = []
        # tmp.append(sheet.cell(row=i, column=2).value) #name
        # tmp.append(sheet.cell(row=i, column=1).value) #recipe id

        # tmp_str=(sheet.cell(row=i, column=3).value) #cleaned ingredients - need to convert to list
        # for character in "[]'":
        #     tmp_str = tmp_str.replace(character, '')
        # tmp1 = tmp_str.split(',')
        # tmp.append(tmp1)
        
        # tmp_str = (sheet.cell(row=i, column=5).value)
        # for character in "[]":
        #     tmp_str = tmp_str.replace(character, '')
        # tmp1 = tmp_str.split("',")
        # for k in range(0, len(tmp1)):
        #     tmp1[k]=tmp1[k].replace('\'', '')
        # tmp.append(tmp1)
        # # for character in "[]":
        # #     tmp_str = tmp_str.replace(character, '')
        
        # # tmp.append(sheet.cell(row=i, column=5).value) #directions - need to convert to list
        # data.append(tmp)
        # print("added")

        valid += 1

        if(valid==2000):
            break
        elif (valid % 100 == 0):
            print(valid)

    except:
        continue




# print(data[0])
print(data[1])