import os
import openpyxl
import torch
from PIL import Image
import torchvision.transforms as transforms


loc = ("excel_datasheet.xlsx")

wb = openpyxl.load_workbook(loc, read_only=True)
sheet = wb.active


valid_ids = {} #change this to dict
for root, dirs, files in os.walk(r"C:\Users\angel\OneDrive\Desktop\Github Repositories\aps360-recipe-ai\allrecipes\test_images"):
    for filename in files:
        # valid_ids.append(filename[:-4])
        valid_ids[int(filename[:-4])] = filename
        
print("Valid ID Collection done") #30 seconds

os.chdir(r"C:\Users\angel\OneDrive\Desktop\Github Repositories\aps360-recipe-ai\allrecipes\test_images")

# print(valid_ids[1])

# # print(valid_ids)
data = []
transform = transforms.Compose([transforms.PILToTensor()])
for i in range(2, 5000):
    # if (sheet.cell(row=i, column=1).value in valid_ids):
    try:
        img_name = valid_ids[sheet.cell(row=i, column=1).value]
        #add data to main test data list

        tmp = []
        tmp1 = []

        image = Image.open(img_name)
        tmp_tensor = transform(image)
        tmp.append(tmp_tensor)

        
        # tmp.append(sheet.cell(row=i, column=2).value) #name
        # tmp.append(sheet.cell(row=i, column=1).value) #recipe id

        #adding cleaned ingredients
        tmp_str=(sheet.cell(row=i, column=3).value)
        for character in "[]'":
            tmp_str = tmp_str.replace(character, '')
        tmp1 = tmp_str.split(',')
        tmp.append(tmp1)
        
        # #cleaning cooking directions; uncomment this block if need instructions
        # tmp_str = (sheet.cell(row=i, column=5).value)
        # for character in "[]":
        #     tmp_str = tmp_str.replace(character, '')
        # tmp1 = tmp_str.split("',")
        # for k in range(0, len(tmp1)):
        #     tmp1[k]=tmp1[k].replace('\'', '')
        # tmp.append(tmp1)
        # for character in "[]":
        #     tmp_str = tmp_str.replace(character, '')
        
        # tmp.append(sheet.cell(row=i, column=5).value) #directions - need to convert to list
        data.append(tmp)
        # print("added")
    except:
        continue
    
    
os.chdir(r"C:\Users\angel\OneDrive\Desktop\Github Repositories\aps360-recipe-ai\allrecipes")


print("Total test datapoints: {}".format(len(data)))
# print(data)